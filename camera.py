import os
import time
import cv2
from ultralytics import YOLO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from system_monitor import start_monitoring, stop_monitoring
from picamera2 import Picamera2

# Настройки путей
MODEL = "models/yolov8n.pt"
SYSTEM_METRICS_FILE = f"share/system_metrics.csv"
EXCEL_FILE = f"share/results_r5/camera_results_{MODEL.split('.')[0]}.xlsx"
VIDEO_OUTPUT = f"share/save_videos_from_camera/output_video_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4"
FRAMES_DIR = f"share/save_frames_from_camera"
NUMBER_OF_FRAMES = 10

# Создаем директории, если их нет
os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
os.makedirs(os.path.dirname(VIDEO_OUTPUT), exist_ok=True)

def init_file():
    """Создает новый Excel-файл с шапкой"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detection Results"
    
    headers = [
        'Timestamp', 'Frame', 'Inference Time (s)',
        'Detected Objects', 'Confidence'
    ]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws.column_dimensions[col_letter].width = 20
    
    wb.save(EXCEL_FILE)

def write_to_file(data):
    """Добавляет данные в Excel-файл"""
    from openpyxl import load_workbook
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

def process_frame(model, frame, frame_count):
    """Обрабатывает кадр и возвращает данные для Excel"""
    # Детекция
    start_time = time.time()
    results = model(frame)
    inference_time = time.time() - start_time

    # Сохранение изображения с bbox
    result_img = results[0].plot()
    frame_filename = os.path.join(FRAMES_DIR, f"frame_{frame_count:04d}.jpg")
    cv2.imwrite(frame_filename, result_img)

    # Визуализация результатов
    annotated_frame = results[0].plot()

    # Детекция объектов
    detected = []
    confs = []
    for box in results[0].boxes:
        class_id = int(box.cls.item())
        confidence = float(box.conf.item())
        detected.append(model.names[class_id])
        confs.append(f"{confidence:.2f}")

    return [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        frame_count,
        float(f"{inference_time:.3f}"),
        ", ".join(detected) if detected else "None",
        ", ".join(confs) if confs else "0"
    ], annotated_frame

def main():
    # Инициализация мониторинга системы
    print("Запуск мониторинга системы...")
    queue, monitor_process = start_monitoring(SYSTEM_METRICS_FILE)
    
    try:
        # Инициализация Excel
        init_file()
        model = YOLO(MODEL)
        
        # Инициализация камеры
        picam2 = Picamera2()
        preview_config = picam2.create_preview_configuration(main={"size": (1280, 720)})
        picam2.configure(preview_config)
        picam2.start()
        
        # Получаем параметры видео для сохранения
        frame_width = 1280
        frame_height = 720
        fps = 30  # Может потребоваться корректировка
        
        # Инициализация VideoWriter
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(VIDEO_OUTPUT, fourcc, fps, (frame_width, frame_height))
        
        print("Начало обработки видеопотока с камеры...")
        print(f"Видео сохраняется в: {VIDEO_OUTPUT}")
        print("Нажмите 'q' для выхода")
        
        start_time = time.time()
        frame_count = 0
        
        while True:
            # Получаем кадр с камеры
            frame = picam2.capture_array()
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
            
            # Обрабатываем кадр
            frame_count += 1
            data, annotated_frame = process_frame(model, frame_rgb, frame_count)
            
            # Записываем данные в Excel (например, каждый 10-й кадр)
            if frame_count % 10 == 0:
                write_to_file(data)
            
            # Показываем результат
            cv2.imshow("Raspberry Pi Camera - YOLOv8 Detection", frame_rgb)
            
            # Выход по нажатию 'q'
            if frame_count == NUMBER_OF_FRAMES:
                break
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        
        final_time = time.time() - start_time
        
        print(f"Обработка завершена! Результаты сохранены:")
        print(f"- Видео: {VIDEO_OUTPUT}")
        print(f"- Данные детекции: {EXCEL_FILE}")
        print(f"Всего обработано кадров: {frame_count}")
        
    finally:
        # Закрываем видеофайл
        if 'out' in locals():
            out.release()
        
        # Остановка камеры
        picam2.stop()
        cv2.destroyAllWindows()
        
        # Остановка мониторинга и вывод статистики
        print("\nОстановка мониторинга системы...")
        from system_monitor import SystemMonitor
        monitor = SystemMonitor()
        stop_monitoring(queue, monitor_process)
        
        # Чтение и анализ собранных метрик
        import pandas as pd
        try:
            df = pd.read_csv(SYSTEM_METRICS_FILE)
            if not df.empty:
                cpu_avg = df['cpu_usage'].mean()
                ram_avg = df['ram_usage'].mean()
                temp_avg = df[df['cpu_temp'] >= 0]['cpu_temp'].mean()
                
                print("\nСредние значения за период работы:")
                print(f"CPU usage: {cpu_avg:.2f}%")
                print(f"RAM usage: {ram_avg:.2f} MB")
                print(f"CPU temp: {temp_avg:.1f}°C" if temp_avg >= 0 else "CPU temp: N/A")
                print(f"Total time: {final_time:.2f} seconds")
                print(f"FPS: {frame_count/final_time:.2f}")
        except Exception as e:
            print(f"Ошибка при анализе метрик: {e}")

if __name__ == "__main__":
    main()
