import os
import time
import cv2
from ultralytics import YOLO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from system_monitor import start_monitoring, stop_monitoring

# Настройки путей
INPUT_DIR = "images"
MODEL = "models/yolov8n.pt"
# OUTPUT_DIR = f"share/results_{MODEL.split('.')[0]}" # Не используется
SYSTEM_METRICS_FILE = f"share/system_metrics.csv"
EXCEL_FILE = f"share/results_r5/images_results_{MODEL.split('.')[0]}.xlsx"

# os.makedirs(OUTPUT_DIR, exist_ok=True)

def init_file():
    """Создает новый Excel-файл с шапкой"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detection Results"
    
    headers = [
        'Timestamp', 'Image Name', 'Inference Time (s)',
        'Detected Objects', 'Confidence'
    ]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws.column_dimensions[col_letter].width = 20
    
    wb.save(EXCEL_FILE)

def write_to_file(data):
    """Добавляет данные в Excel-файл"""
    from openpyxl import load_workbook
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

# def process_image(model, img_path, output_dir):
def process_image(model, img_path):
    """Обрабатывает изображение и возвращает данные для Excel"""
    img = cv2.imread(img_path)
    if img is None:
        print(f"Ошибка загрузки: {img_path}")
        return None

    # Детекция
    start_time = time.time()
    results = model(img)
    inference_time = time.time() - start_time

    # Сохранение изображения с bbox
    base_name = os.path.basename(img_path)
    # output_path = os.path.join(output_dir, f"result_{base_name}")
    # result_img = results[0].plot()
    # cv2.imwrite(output_path, result_img)

    # Детекция объектов
    detected = []
    confs = []
    for box in results[0].boxes:
        class_id = int(box.cls.item())
        confidence = float(box.conf.item())
        detected.append(model.names[class_id])
        confs.append(f"{confidence:.2f}")

    return [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        base_name,
        float(f"{inference_time:.3f}"),
        ", ".join(detected) if detected else "None",
        ", ".join(confs) if confs else "0"
    ]

def main():

    # Инициализация мониторинга системы
    print("Запуск мониторинга системы...")
    queue, monitor_process = start_monitoring(SYSTEM_METRICS_FILE)
    
    try:
        # Инициализация Excel
        init_file()
        model = YOLO(MODEL)
        
        # Получаем все изображения из папки
        image_files = [f for f in os.listdir(INPUT_DIR) 
                     if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        
        if not image_files:
            print("Нет изображений в папке!")
            return

        print(f"Начало обработки {len(image_files)} изображений...")
        
        start_time = time.time()
        
        for img_file in image_files:
            img_path = os.path.join(INPUT_DIR, img_file)
            print(f"Обработка: {img_file}")
            
            # data = process_image(model, img_path, OUTPUT_DIR)
            data = process_image(model, img_path)
            if data:
                write_to_file(data)
        
        final_time = time.time() - start_time
        
        print(f"Обработка завершена! Результаты сохранены в {EXCEL_FILE}")
        
    finally:
        # Остановка мониторинга и вывод статистики
        print("\nОстановка мониторинга системы...")
        from system_monitor import SystemMonitor
        monitor = SystemMonitor()
        stop_monitoring(queue, monitor_process)
        
        # Чтение и анализ собранных метрик
        import pandas as pd
        try:
            df = pd.read_csv(SYSTEM_METRICS_FILE)
            if not df.empty:
                cpu_avg = df['cpu_usage'].mean()
                ram_avg = df['ram_usage'].mean()
                temp_avg = df[df['cpu_temp'] >= 0]['cpu_temp'].mean()
                
                print("\nСредние значения за период работы:")
                print(f"CPU usage: {cpu_avg:.2f}%")
                print(f"RAM usage: {ram_avg:.2f} MB")
                print(f"CPU temp: {temp_avg:.1f}°C" if temp_avg >= 0 else "CPU temp: N/A")
                print(f"Time: {final_time}")
        except Exception as e:
            print(f"Ошибка при анализе метрик: {e}")

if __name__ == "__main__":
    main()
